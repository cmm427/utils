# -*- coding: gbk -*

from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, GradientFill, Alignment
from openpyxl.chart import AreaChart, BarChart, Reference, Series
from openpyxl.utils import get_column_letter
import collections
import configparser
import codecs

base = [chr(x) for x in range(ord('A'), ord('A') + 26)]


def dec2tsix(num):
    l = []
    if num < 0:
        return '-' + dec2tsix(abs(num))
    while True:
        num, rem = divmod(num, 26)
        l.append(base[rem - 1])
        if num == 0:
            return ''.join(l[::-1])


def create_area_chart(ws, chart=AreaChart()):
    chart.style = 13

    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    data = Reference(ws, min_col=2, min_row=1, max_col=ws.max_column, max_row=ws.max_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    col = dec2tsix(ws.max_column + 2)
    chart_cell = '%s%s' % (col, ws.max_row + 2)

    ws.add_chart(chart, chart_cell)


def create_bar_chart(work_sheet, bar_chart=BarChart()):
    bar_chart.style = 13

    data = Reference(work_sheet, min_col=2, min_row=1, max_col=work_sheet.max_column, max_row=work_sheet.max_row)
    bar_chart.add_data(data, titles_from_data=True)

    cats = Reference(work_sheet, min_col=1, min_row=2, max_row=work_sheet.max_row)
    bar_chart.set_categories(cats)

    cl = get_column_letter(work_sheet.max_column + 2)
    chart_cell = '%s%s' % (cl, work_sheet.max_row + 2)

    work_sheet.add_chart(bar_chart, chart_cell)


def style_range(ws, cell_range, font=None, fill=None, border=Border(), alignment=None):
    rows = ws[cell_range]

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    for row in rows:
        for cell in row:
            cell.border = border
            cell.alignment = alignment

        if row == rows[0]:
            for cell in row:
                cell.font = Font(bold=False, color='000000')
                cell.fill = PatternFill("solid", fgColor="B8CCE4")

        else:
            for cell in row:
                cell.font = font
                cell.fill = fill
                if cell.column == 'B':
                    cell.alignment = Alignment(horizontal="general", vertical="center")


cfg = configparser.ConfigParser()
cfg.read_file(codecs.open("config.ini", "r", "utf8"))

srcfile = cfg.get("directory","source")
targetfile = cfg.get("directory", "target")

title = str(cfg.get("excel", "title")).split(",")
chart_title = str(cfg.get("excel", "chart_title")).split(",")

wb = load_workbook(srcfile)
ws = wb.active
ws_result = wb.create_sheet("result")
ws_chart = wb.create_sheet("chart")

srcidx = []
chartidx = []
chart1 = []
targetidx = ['A', 'B', 'C', 'D', 'E']

# get column index
for cell in ws['1']:
    if cell.value in title:
        srcidx.append(cell.column)

# copy columns to result sheet
for i in range(1, len(srcidx) + 1):
    for r in range(1, ws.max_row + 1):
        si = '%s%s' % (srcidx[i - 1], r)
        ti = '%s%s' % (targetidx[i - 1], r)
        ws_result[ti].value = ws[si].value

# count chart
for cell in ws['1']:
    if cell.value in chart_title:
        chartidx.append(cell.column)

for i in range(1, len(chartidx) + 1):
    tmp = []
    for r in range(2, ws.max_row + 1):
        idx = '%s%s' % (chartidx[i - 1], r)
        if ws[idx].value != '0000-00-00':
            tmp.append(ws[idx].value)
    chart1.append(tmp)

# count
s = set()
chart2 = []
for c in chart1:
    s = s | set(c)
    chart2.append(dict(collections.Counter(c).items()))

ss = sorted(s)
chart_list = []
for d in ss:
    tmp = []
    tmp.append(d)

    for c in chart2:
        if c.get(d) is not None:
            tmp.append(c.get(d))
        else:
            tmp.append('0')

    chart_list.append(tmp)

chart_list.insert(0, [u"日期", u"提出数", u"关闭数"])

# copy chart_list to chart sheet
for row in chart_list:
    ws_chart.append(row)

# create chart
chart = BarChart()
create_bar_chart(ws_chart, chart)

# style worksheet
side = Side(border_style='thin', color='000000')
border = Border(left=side, right=side, top=side, bottom=side)
font = Font(bold=False, color='000000')
fill = PatternFill("solid", fgColor="FFFFFF")
ws_result.column_dimensions['B'].width = 100
al = Alignment(horizontal="center", vertical="center")

col = dec2tsix(ws_result.max_column)
bottom_right = '%s%s' % (col, ws_result.max_row)
cell_range = "A1:" + bottom_right

style_range(ws_result, cell_range, font=font, fill=fill, border=border, alignment=al)

# delete ws
wb.remove(ws)

wb.save(targetfile)
